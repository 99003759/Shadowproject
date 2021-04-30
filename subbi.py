import openpyxl
from datetime import datetime
from openpyxl import Workbook
wb = openpyxl.load_workbook('subbixl.xlsx')
sheets = ['Sheet1']
xin = input("P or A")
count = 0
yin = str(input("Date"))
my = datetime.strptime(yin, "%d-%m-%Y")
for sheet in sheets:
    sh = wb[sheet]  # Get a sheet from the workbook.
    max_r = sh.max_row
    max_c = sh.max_column
    for r in range(1, max_r+1):
        for c in range(4, max_c+1):
            if sh.cell(row=1, column=c).value == my:
                if sh.cell(row=r, column=c).value == xin:
                    count = count + 1
print(str((count/(max_r-1))*100)+"%")
